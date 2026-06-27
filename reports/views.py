import io
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Q, Sum
from django.utils import timezone

from core.models import Asset, AuditLog, Department, AssetCategory, Transaction


def _get_ip(request):
    x = request.META.get('HTTP_X_FORWARDED_FOR')
    return x.split(',')[0].strip() if x else request.META.get('REMOTE_ADDR', '')


def _filter_assets(request):
    qs = Asset.objects.select_related('category', 'department', 'custodian')
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    department = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if q:
        qs = qs.filter(Q(asset_id__icontains=q) | Q(name__icontains=q))
    if status:
        qs = qs.filter(status=status)
    if category:
        qs = qs.filter(category_id=category)
    if department:
        qs = qs.filter(department_id=department)
    if date_from:
        qs = qs.filter(purchase_date__gte=date_from)
    if date_to:
        qs = qs.filter(purchase_date__lte=date_to)
    return qs


@login_required
def stock_report(request):
    qs = _filter_assets(request).order_by('-created_at')
    if request.user.role == 'hod' and request.user.department:
        qs = qs.filter(department=request.user.department)
    total_value = qs.aggregate(v=Sum('purchase_cost'))['v'] or 0

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'total_value': total_value,
        'categories': AssetCategory.objects.all(),
        'departments': Department.objects.all(),
        'status_choices': Asset.STATUS_CHOICES,
        'current_filters': {
            'q': request.GET.get('q', ''),
            'status': request.GET.get('status', ''),
            'category': request.GET.get('category', ''),
            'department': request.GET.get('department', ''),
            'date_from': request.GET.get('date_from', ''),
            'date_to': request.GET.get('date_to', ''),
        },
    }
    return render(request, 'reports/stock.html', context)


@login_required
def audit_log(request):
    if request.user.role not in ('superadmin', 'auditor'):
        raise PermissionDenied
    qs = AuditLog.objects.select_related('user').order_by('-timestamp')
    q = request.GET.get('q', '').strip()
    user_q = request.GET.get('user', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if q:
        qs = qs.filter(action__icontains=q)
    if user_q:
        qs = qs.filter(Q(user__username__icontains=user_q) | Q(user__first_name__icontains=user_q))
    if date_from:
        qs = qs.filter(timestamp__date__gte=date_from)
    if date_to:
        qs = qs.filter(timestamp__date__lte=date_to)

    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'reports/audit_log.html', {
        'page_obj': page,
        'current_filters': {'q': q, 'user': user_q, 'date_from': date_from, 'date_to': date_to},
    })


@login_required
def reconciliation(request):
    if request.user.role not in ('superadmin', 'auditor', 'store_officer'):
        raise PermissionDenied
    asset = None
    discrepancy = None
    barcode = request.GET.get('barcode', '').strip()

    if barcode:
        try:
            asset = Asset.objects.select_related('department', 'custodian', 'category').get(
                Q(barcode_number=barcode) | Q(asset_id=barcode)
            )
        except Asset.DoesNotExist:
            pass

    if request.method == 'POST' and asset:
        note = request.POST.get('discrepancy_note', '').strip()
        if note:
            AuditLog.objects.create(
                user=request.user,
                action=f'DISCREPANCY flagged for {asset.asset_id}: {note}',
                model_name='Asset',
                object_id=str(asset.pk),
                ip_address=_get_ip(request),
            )
            discrepancy = 'flagged'

    return render(request, 'reports/reconciliation.html', {
        'asset': asset,
        'barcode': barcode,
        'discrepancy': discrepancy,
    })


@login_required
def export_centre(request):
    if request.user.role not in ('superadmin', 'auditor', 'store_officer', 'bursar'):
        raise PermissionDenied
    return render(request, 'reports/export_centre.html', {
        'categories': AssetCategory.objects.all(),
        'departments': Department.objects.all(),
        'status_choices': Asset.STATUS_CHOICES,
    })


@login_required
def export_excel(request):
    if request.user.role not in ('superadmin', 'auditor', 'store_officer', 'bursar'):
        raise PermissionDenied
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    qs = _filter_assets(request).order_by('asset_id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Report'

    headers = ['Asset ID', 'Name', 'Category', 'Department', 'Status', 'Condition',
               'Custodian', 'Serial No.', 'Purchase Date', 'Purchase Cost (₦)', 'Location']
    header_fill = PatternFill(start_color='1a1d27', end_color='1a1d27', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A2'

    for row, asset in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=asset.asset_id)
        ws.cell(row=row, column=2, value=asset.name)
        ws.cell(row=row, column=3, value=asset.category.name)
        ws.cell(row=row, column=4, value=asset.department.name)
        ws.cell(row=row, column=5, value=asset.get_status_display())
        ws.cell(row=row, column=6, value=asset.get_condition_display())
        ws.cell(row=row, column=7, value=asset.custodian.get_full_name() if asset.custodian else '')
        ws.cell(row=row, column=8, value=asset.serial_number)
        ws.cell(row=row, column=9, value=str(asset.purchase_date) if asset.purchase_date else '')
        ws.cell(row=row, column=10, value=float(asset.purchase_cost) if asset.purchase_cost else '')
        ws.cell(row=row, column=11, value=asset.location_description)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="lasu_ims_stock_{ts}.xlsx"'
    return response


@login_required
def export_pdf(request):
    if request.user.role not in ('superadmin', 'auditor', 'store_officer', 'bursar'):
        raise PermissionDenied
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse('xhtml2pdf not installed.', status=500)

    qs = _filter_assets(request).order_by('asset_id')
    total_value = qs.aggregate(v=Sum('purchase_cost'))['v'] or 0

    html = render(request, 'reports/pdf_template.html', {
        'assets': qs,
        'total_value': total_value,
        'generated_at': timezone.now(),
    }).content.decode('utf-8')

    buf = io.BytesIO()
    pisa.CreatePDF(html, dest=buf)
    buf.seek(0)

    ts = timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lasu_ims_stock_{ts}.pdf"'
    return response
